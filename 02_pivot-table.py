import pandas as pd

df = pd.read_excel('supermarket_sales.xlsx')

# How much is each gender spending per product line?
# df[] selects one column, df[[]] selects multiple
df = df[['Gender', 'Product line', 'Total']]


pivot_table = df.pivot_table(index="Gender", columns="Product line", values="Total",
               aggfunc="sum").round(0)

#print(pivot_table)

'''
Product line  Electronic accessories  Fashion accessories  Food and beverages  Health and beauty  Home and lifestyle  Sports and travel
Gender
Female                       27102.0              30437.0             33171.0            18561.0             30037.0            28575.0
Male                         27236.0              23868.0             22974.0            30633.0             23825.0            26548.0

This is cool!
'''

pivot_table.to_excel('pivot_table.xlsx', 'Report', startrow=4)

# Bar chart for the pivot table
# Using openpyxl, my old friend <3

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Pivot table is located from A5 to G7

min_column = wb.active.min_column # 1
max_column = wb.active.max_column # 7
min_row = wb.active.min_row # 5
max_row = wb.active.max_row # 7

barchart = BarChart()
data = Reference(sheet,
          min_col=min_column+1, #Pivot table starts in A but data starts in B
          max_col=max_column,
          min_row=min_row, # Why are we not increasing this? Row 5 is all labels.
          max_row=max_row)

categories = Reference(sheet,
          min_col=min_column,
          max_col=min_column, # <- DO THIS OR ALL THE DATA ENDS UP ON THE AXIS LABELS (ask me how I know)
          min_row=min_row+1, # Because we're doing it here *shrug*
          max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")

barchart.title = "Sales by Product line and Gender"
barchart.style = 2 # Same colors as in video is 2, not 5

wb.save('barchart.xlsx')

# AUTOMATE EXCEL REPORT -> apply-formula.py